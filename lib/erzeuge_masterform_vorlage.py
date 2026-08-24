# ============================================================
# erzeuge_masterform_vorlage.py
# Erzeugt assets/templates_xlsx/fill_a_masterform_vorlage.xlsx - 1.0
#
# Baut die Vorlage/Testdatei fuer den Fill-a-Masterform-Import
# (siehe masterform_import.py) DIREKT aus dessen Mapping-Tabellen -
# damit die Vorlage nie von den tatsaechlich unterstuetzten Werten
# abweichen kann. Bei Aenderungen an masterform_import.py (neue/
# geaenderte Spalte, neue gueltige Werte) hier einfach neu ausfuehren:
#
#   python lib/erzeuge_masterform_vorlage.py
#
# Die Vorlage dient zwei Zwecken:
#   1. Zum Testen von /masterform im Web-Editor, ohne auf eine echte
#      Datei vom anderen Team angewiesen zu sein.
#   2. Als Schema-Referenz fuer die Abstimmung mit "Fill-a-Masterform"
#      (Kopfzeile + Zellkommentare mit den jeweils gueltigen Werten).
# ============================================================

import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill

import masterform_import as mi

ZIEL_PFAD = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "assets", "templates_xlsx", "fill_a_masterform_vorlage.xlsx",
)

# Frei erfundene, aber in sich konsistente Beispielzeile - zeigt ein
# plausibles Wertemuster je Spalte.
BEISPIELZEILE = {
    "excel_zeile": 2, "mlcs_id": "9999", "bereich": "ICF Pharma", "gebaeude": "H600",
    "dok_version": "1", "dok_nummer": None, "systemname": "Beispielsystem XY",
    "anlage": "Beispielanlage", "raum": None,
    "kurzbeschreibung": "Kurzbeschreibung des Systems und seines Verwendungszwecks.",
    "sw_name": "TIA-Portal", "sw_version": "V15.0", "sw_hersteller": "Siemens",
    "hersteller": "Siemens", "lieferantennummer": "OOZ00000000000",
    "gxp_relevant": "ja", "gxp_kritikalitaet": "Major",
    "systemtyp_cis": False, "systemtyp_ce": True, "subtyp": "PCS", "vnap_stufe": None,
    "doku_status": "offen", "gamp_kategorie": "4", "eres_typ": "Typ 2",
    "testtiefe": "mittel", "zone_stufe": "Z2S2", "business_critical": "nein",
    "dokumentart": "Neuerstellung", "geraetekategorie": None, "vq_nvq": "NVQ",
    "qualifizierung_erforderlich": True, "validierung_erforderlich": True,
    "ki_einstufung": "N/A", "subtyp_mehrfach_markiert": None,
    "geraetekategorie_mehrfach_markiert": None,
}

# Spalte -> Liste der laut masterform_import.py gueltigen Werte
# (fuer die Zellkommentare in der Kopfzeile). Spalten ohne Eintrag
# hier sind reine 1:1-Textfelder (siehe PLAIN_FELD_MAPPING) - kein
# fester Wertevorrat, deshalb kein Kommentar noetig.
GUELTIGE_WERTE = {
    "gxp_relevant":                        list(mi.GXP_RELEVANZ),
    "gxp_kritikalitaet":                   list(mi.GXP_KRITIKALITAET),
    "systemtyp_cis":                       ["True", "False"],
    "systemtyp_ce":                        ["True/False (nur informativ, wird aus subtyp abgeleitet)"],
    "subtyp":                              list(mi.SUBTYP),
    "vnap_stufe":                          list(mi.VNAP_STUFE),
    "subtyp_mehrfach_markiert":            ["True = CS-Typ nicht eindeutig, wird NICHT automatisch gesetzt"],
    "gamp_kategorie":                      list(mi.GAMP_KATEGORIE),
    "eres_typ":                            list(mi.ERES_TYP),
    "testtiefe":                           list(mi.TESTTIEFE),
    "business_critical":                  list(mi.BUSINESS_CRITICAL),
    "dokumentart":                         list(mi.DOKUMENTART),
    "geraetekategorie":                    list(mi.GERAETEKATEGORIE),
    "geraetekategorie_mehrfach_markiert":  ["True = Gerätekategorie nicht eindeutig, wird NICHT automatisch gesetzt"],
    "vq_nvq":                              list(mi.VQ_NVQ),
    "ki_einstufung":                       list(mi.KI_EINSTUFUNG),
    "zone_stufe":                          mi.ZONE_FELDER,
    "doku_status":                         ["frei/informativ - wird NICHT automatisch einer Checkbox zugeordnet"],
    "qualifizierung_erforderlich":         ["True/False - wird NICHT automatisch einer Checkbox zugeordnet"],
    "validierung_erforderlich":            ["True/False - wird NICHT automatisch einer Checkbox zugeordnet"],
}

def erzeugen(ziel_pfad=ZIEL_PFAD):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MasterDB_bereinigt"

    header = mi.ERWARTETE_SPALTEN
    ws.append(header)
    ws.append([BEISPIELZEILE.get(spalte) for spalte in header])

    fett = Font(bold=True)
    fuellung = PatternFill("solid", fgColor="DDEBF7")
    for col_idx, name in enumerate(header, start=1):
        zelle = ws.cell(row=1, column=col_idx)
        zelle.font = fett
        zelle.fill = fuellung
        ws.column_dimensions[zelle.column_letter].width = max(14, len(name) + 2)
        werte = GUELTIGE_WERTE.get(name)
        if werte:
            text = "Gültige Werte:\n- " + "\n- ".join(str(w) for w in werte)
            kommentar = Comment(text, "SysBew_Extraktor")
            kommentar.width, kommentar.height = 260, 120
            zelle.comment = kommentar

    ws.freeze_panes = "A2"
    os.makedirs(os.path.dirname(ziel_pfad), exist_ok=True)
    wb.save(ziel_pfad)
    return ziel_pfad

if __name__ == "__main__":
    pfad = erzeugen()
    print(f"✅ Vorlage erzeugt: {pfad}")
